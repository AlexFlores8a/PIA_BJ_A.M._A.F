import json
import statistics as stats
from collections import Counter
from openpyxl import Workbook
import os

def main():
    # Cargar datos multi empresa
    with open("data/clean/multi_clean.json", "r") as f:
        multi = json.load(f)

    # Crear carpeta de salida para Excel
    os.makedirs("results/excel", exist_ok=True)

    # Crear libro de Excel
    wb = Workbook()

    # --- Hoja 1: Datos limpios ---
    ws1 = wb.active
    ws1.title = "Datos_limpios"
    headers = ["Símbolo", "Nombre", "Exchange", "Precio actual", "Precio anterior",
               "Cambio", "Cambio %", "Mínimo día", "Máximo día", "Volumen", "Fecha consulta"]
    ws1.append(headers)
    for item in multi:
        ws1.append([
            item["simbolo"],
            item["nombre"],
            item["exchange"],
            item["precio_actual"],
            item["precio_anterior"],
            item["cambio"],
            item["cambio_porcentaje"],
            item["min_dia"],
            item["max_dia"],
            item["volumen"],
            item["fecha_consulta"]
        ])

    # --- Hoja 2: Estadísticas ---
    ws2 = wb.create_sheet("Estadisticas")
    # Extraer cambios porcentuales (ignorar None)
    cambios = [item["cambio_porcentaje"] for item in multi if item["cambio_porcentaje"] is not None]
    if cambios:
        media = stats.mean(cambios)
        mediana = stats.median(cambios)
        try:
            moda = stats.mode(cambios)
        except stats.StatisticsError:
            moda = "No hay moda única"
        rango_val = max(cambios) - min(cambios)
        minimo = min(cambios)
        maximo = max(cambios)
        desviacion = stats.stdev(cambios) if len(cambios) > 1 else 0

        ws2.append(["Métrica", "Valor"])
        ws2.append(["Media (%)", media])
        ws2.append(["Mediana (%)", mediana])
        ws2.append(["Moda", moda])
        ws2.append(["Rango (%)", rango_val])
        ws2.append(["Mínimo (%)", minimo])
        ws2.append(["Máximo (%)", maximo])
        ws2.append(["Desviación estándar (%)", desviacion])
    else:
        ws2.append(["No hay datos suficientes para estadísticas"])

    # --- Hoja 3: Tabla de frecuencias (por exchange) ---
    ws3 = wb.create_sheet("Frecuencias")
    exchanges = [item["exchange"] for item in multi]
    frecuencia = Counter(exchanges)
    ws3.append(["Exchange", "Frecuencia"])
    for exch, count in frecuencia.items():
        ws3.append([exch, count])

    # Guardar el archivo Excel en la ruta correcta
    ruta_excel = "results/excel/reporte_final.xlsx"
    wb.save(ruta_excel)
    print(f"✅ Archivo Excel guardado en {ruta_excel}")
    print("   Hojas creadas: Datos_limpios, Estadisticas, Frecuencias")

if __name__ == "__main__":
    main()